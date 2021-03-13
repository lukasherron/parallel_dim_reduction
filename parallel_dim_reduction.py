#!/usr/bin/env python
from sklearn.decomposition import PCA
from scipy.io import loadmat
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import openpyxl as pxl
from sklearn.model_selection import train_test_split
from scipy.ndimage import gaussian_filter1d
import os
import glob


def KL_divergence(arr_1, arr_2):
    # arr_1 is data, arr_2 is theory
    temp = np.ma.log(np.divide(arr_1, arr_2, where=arr_2 != 0).astype('float64'))
    temp = temp.filled(0)
    temp[np.isnan(temp)] = 0
    KL = np.nansum(np.nansum(np.multiply(arr_1, temp)))
    return KL

#excel definitions
def excel_cell(row, col):
    """ Convert given row and column number to an Excel-style cell name. """
    LETTERS = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    result = []
    while col > 0:
        col, rem = divmod(col - 1, 26)
        result[:0] = LETTERS[rem]
    return ''.join(result) + str(row)

def load_sheet_rows(sheet):
    # constructing dictionary from sheet
    keys = []
    if sheet.title == "params" or sheet.title == "eval":
        m = 1
        while m:
            key = sheet["A" + str(m)].value
            if key is None:
                break
            else:
                keys.append(key)
                m += 1
        new_dict = {key: None for key in keys}
    else:
        keys.append(sheet.title)

    # getting data corresponding to keys in new_dict
    if sheet.title == "params" or sheet.title == "eval":
        for idx, key in list(enumerate(new_dict.keys())):
            temp, k = [], 1
            while k:
                entry = sheet[excel_cell(idx+1, k+1)].value
                if entry is None:
                    break
                else:
                    temp.append(entry)
                    k += 1
            new_dict[key] = temp
            if len(temp) == 1:
                [temp] = temp
                new_dict[key] = temp
    else:
        new_dict = {key: pd.DataFrame(sheet.values).to_numpy() for key in keys}

    return new_dict

def create_filetree(path_to_data_dir, path_to_dataset, model_type, further_spec=None):

    parent_dir = lambda: os.chdir(os.path.dirname(os.getcwd()))

    def conditional(directory):
        try:
            os.chdir(directory)
        except:
            os.mkdir(directory)
            os.chdir(directory)

    cd = os.getcwd()
    os.chdir(path_to_data_dir)
    f = os.path.basename(path_to_dataset)
    (name, ext) = os.path.splitext(f)
    conditional(name)
    conditional(model_type)
    if further_spec is not None:
        conditional(further_spec)
    arr = ["models", "predictions", "plots_and_figs"]
    for path in arr:
        conditional(path)
        for k in range(6, 29):
            conditional("nK=" + str(k))
            for l in range(0, 11):
                conditional("nL=" + str(l))
                parent_dir()
            parent_dir()
        parent_dir()
    os.chdir(cd)
    
    return name

class Parallel_Dimensionality_Reduction(object):

    def __init__(self, param_dict, data_dict, eval_dict, grad_dict):

        self.param_dict = param_dict
        self.data_dict = data_dict
        self.eval_dict = eval_dict
        self.grad_dict = grad_dict

    def eval_Q(self):

        Z = self.data_dict["Z"]
        thet = self.data_dict["thet"]

        Q = np.exp(-Z @ thet)
        Q[np.isnan(Q)] = 0
        norms = Q.sum(1)
        Q = np.divide(Q.T, norms, where=norms != 0).T

        self.data_dict["Q"] = Q

    def gradient_update(self):

        Q = self.data_dict["Q"]
        Z = self.data_dict["Z"]
        thet = self.data_dict["thet"]
        C = self.data_dict["C"]
        xs = self.data_dict["xs_train"]
        metz = self.data_dict["metz_train"]

        alph = self.param_dict["alph"]
        nPC = self.param_dict["nPC"]
        etaZ = self.param_dict["etaZ"]
        etaT = self.param_dict["etaT"]
        etaC = self.param_dict["etaC"]

        delt = xs - Q
        Zcon = Z[:, 0:nPC]
        grthet = (1 - alph) * Z.T @ delt
        grz = (1 - alph) * delt @ thet.T
        # gradient for free components

        grz_con = -2 * alph * (metz - Zcon @ C) @ C.T
        grz[:, 0:nPC] += grz_con
        grc = -2 * alph * Zcon.T @ (metz - Zcon @ C)
        Z -= etaZ * grz
        thet -= etaT * grthet
        C -= etaC * grc

        self.data_dict["Z"] = Z
        self.data_dict["thet"] = thet

        self.grad_dict["grz"] = grz
        self.grad_dict["grthet"] = grthet
        self.grad_dict["grc"] = grc

    def eval_training_performance(self):

        metz = self.data_dict["metz_train"]
        Zcon = self.data_dict["Zcon"]
        C = self.data_dict["C"]
        xs_train = self.data_dict["xs_train"]
        Q = self.data_dict["Q"]
        alph = self.param_dict["alph"]
        idx_train =self.data_dict["idx_train"]

        e1 = np.linalg.norm(metz - Zcon @ C)
        e1 = alph * e1 ** 2

        KL_train = KL_divergence(xs_train, Q)/len(idx_train)
        e2 = (1 - alph) * KL_train

        self.eval_dict["e1"] = e1
        self.eval_dict["e2"] = e2
        self.eval_dict["KL_train"].append(KL_train)

    def step(self):

        self.eval_Q()
        self.gradient_update()

    def gradient_descent(self, path, samp_freq, stop_search, maxiter, validate=True, verbose=True):
        i = 0
        check_stop = False

        while i < maxiter:
            self.step()
            i += 1
            if i % samp_freq == 0:
                self.eval_training_performance()
                if validate:
                    PFM = Predict_From_Model()
                    PFM.predict_microbiome(self.param_dict, self.data_dict, validate=validate)
                    self.eval_dict["KL_val"].append(PFM.pred_dict["KL"])
                    KL_1 = np.array(self.eval_dict["KL_val"])
                    KL_2 = np.array(self.eval_dict["KL_train"])
#                     KL_tot = (KL_1 + KL_2)[int((stop_search - samp_freq) / samp_freq):]
                    KL_tot = (KL_1)[int((stop_search - samp_freq) / samp_freq):]
                # writer = DataWriter()
                # writer.write_model_workbook(path, self.param_dict, self.eval_dict, self.data_dict)

                # Detecting drop in KL divergence that corresponds to optimal generality (after 55,000 iterations)
#                 if i > stop_search + 2*samp_freq:
#                     KL_p = gaussian_filter1d(KL_tot[1:-1], 5) - gaussian_filter1d(KL_tot[0:-2], 5)  # smoothing noise
#                     print("KL_prime =",'%.3f' % KL_p[-1])
#                     print("KL = ", KL_tot[-1])
# #                     if KL_p[-1] < -3.5:
#                     if KL_p[-1] < -40:
#                         check_stop = True
#                     if check_stop is True and KL_p[-1] >= 0:
#                         break
                    if i > 20000:
                        break

        self.eval_training_performance()
        if validate:
            PFM = Predict_From_Model()
            PFM.predict_microbiome(self.param_dict, self.data_dict, validate=True)
            self.eval_dict["KL_val"].append(PFM.pred_dict["KL"])
        writer = DataWriter()
        writer.write_model_workbook(path, self.param_dict, self.eval_dict, self.data_dict)

    def finalize(self):
        return self.data_dict, self.eval_dict, self.grad_dict


class PCA_reduction(object):

    def __init__(self, param_dict, data_dict, eval_dict):
        self.param_dict = param_dict
        self.data_dict = data_dict
        self.eval_dict = eval_dict

    def dim_reduction(self, guess):
        xs = self.data_dict["xs_train"]
        nPC = self.param_dict["nPC"]
        nL = self.param_dict["nL"]
        nMet = self.param_dict["nMet"]
        metz = self.data_dict["metz_train"]

        [nSamp, _] = xs.shape
        pca = PCA(n_components=nPC)
        Zcon = pca.fit_transform(metz)
        [C, _] = [pca.components_, pca.explained_variance_]
        C = C.T
        C = C[:nPC, :]

        if guess:
            C = np.random.normal(size=(nPC, nMet))
            Zfr = np.random.random(size=(nSamp, nL))
            Z = np.hstack((Zcon, Zfr))

        self.data_dict["Z"] = Z
        self.data_dict["C"] = C
        self.data_dict["Zcon"] = Zcon
        self.eval_dict["ll_pca"] = pca.score(metz)

    def finalize(self):
        return self.data_dict, self.eval_dict


class DataLoader(object):

    def __init__(self):
        self.mat = None
        self.labels = None
        self.nMet = None
        self.mat_keys = None
        self.nSamp = None
        self.mu_met = None
        self.sg_met = None

    def load_data(self, path):
        mat = loadmat(path)
        self.mat_keys = mat.keys()
        self.mat = mat
        xs = mat['xs'].toarray()
        metx = mat["metx"]
        labels = [mat["metnames"][0][i][0] for i in range(len(mat["metnames"][0]))]
        [nSamp, nMet] = metx.shape

        self.nMet = nMet
        self.nSamp = nSamp

        return xs, metx, labels

    def split_data(self, test_size, val_size):
        indices = list(range(self.nSamp))
        if val_size == 0:
            train_indices, test_indices = train_test_split(indices, test_size=test_size)
            val_indices = [[]]
        else:
            train_indices, test_and_val_indices = train_test_split(indices, test_size=test_size + val_size)
            test_indices, val_indices = train_test_split(test_and_val_indices, test_size=val_size/(test_size + val_size))

        return train_indices, test_indices, val_indices


    def preprocess_metadata(self, metx, train_idx, test_idx, val_idx):

        metx_train = metx[train_idx]
        metx_test = metx[test_idx]
        metx_val = metx[val_idx]
        mu_met = np.mean(metx_train, axis=0)
        sg_met = np.diagonal(np.cov(metx_train.T))
        metz_train = np.divide((metx_train - mu_met), sg_met)
        metz_test = np.divide((metx_test - mu_met), sg_met)
        metz_val = np.divide((metx_val - mu_met), sg_met)

        self.mu_met = mu_met
        self.sg_met = sg_met

        return metz_train, metz_test, metz_val
    
    def load_model(self, path):

        wb = pxl.load_workbook(path, read_only=True)
        names = wb.sheetnames[1:]
        k = 0
        if names[0] == "params":
            active_ws = wb[names[0]]
            model_param_dict = load_sheet_rows(active_ws)
            k += 1
        else:
            model_param_dict = {}

        if names[1] == "eval":
            active_ws = wb[names[1]]
            model_eval_dict = load_sheet_rows(active_ws)
            k += 1
        else:
             model_eval_dict = {}
        model_data_dict = {}
        for name in names[k:]:
            active_ws = wb[name]
            data = load_sheet_rows(active_ws)
            model_data_dict[list(data.keys())[0]] = data[list(data.keys())[0]]

        return model_param_dict, model_eval_dict, model_data_dict

    def load_from_params(self, nK, nL, alph, entry, data_type):

        def subdirs(path):
            for entry in os.scandir(path):
                if not entry.name.startswith('.') and entry.is_dir():
                    yield entry.name

        def find_files(path):
            result = [y for x in os.walk(path) for y in glob.glob(os.path.join(x[0], '*.xlsx'))]     
            return result
        
        def pop_path(filename):
            i = 0
            while i != -1:
                i = filename.find("/")
                filename = filename[i+1:]
            return filename
        
        def get_str_from_name(filename, param_name):
            filename = pop_path(filename)
            i = filename.find(param_name)
            j = len(param_name)
            temp = filename[i+j+1:]
            k = temp.find('_')
            param_value = temp[0:k]
            return param_value

        def remove_extra_files(filenames, param_dict):
            for idx, key in list(enumerate(param_dict.keys())):
                if param_dict[key] != 'all':
                    x = param_dict[key]
                    new_filenames = filenames.copy()
                    for name in filenames:
                        param = get_str_from_name(name, key)
                        if param != str(x):
                            new_filenames.remove(name)
                    filenames = new_filenames
            return filenames

        def get_entries(filenames, entry):
            i = 0
            alph_arr, nK_arr, nL_arr, data_arr, param_arr = [], [], [], [], []
            for name in filenames:
                alph = get_str_from_name(name, 'alph')
                alph_arr.append(float(alph))
                nK = get_str_from_name(name, 'nK')
                nK_arr.append(int(nK))
                nL = get_str_from_name(name, 'nL')
                nL_arr.append(int(nL))
                param_arr.append([int(nK), int(nL), float(alph)])

                wb = pxl.load_workbook(name, read_only=True)
                sheetnames = wb.sheetnames

                for sheetname in sheetnames:
                    active_ws = wb[sheetname]
                    dic = load_sheet_rows(active_ws)
                    for key in list(dic.keys()):
                        if key == entry:
                            [[data]] = dic[key]
                data_arr.append(data)
                
                i += 1
                if i % 10 == 0:
                    print("i = ", i)

            alph_set = set(alph_arr)
            nK_set = set(nK_arr)
            nL_set = set(nL_arr)

            return alph_set,  nK_set, nL_set, data_arr, param_arr
            

        param_dict = {"nK": nK, "nL": nL, "alph": alph}
        n_param = 0
        all_entries, particulars = [], []
        for idx, key in list(enumerate(param_dict.keys())):
            if param_dict[key] == "all":
                n_param += 1
                all_entries.append(key)
            else:
                particulars.append(key)

        path_prefix = "/blue/pdixit/lukasherron/parallel_dim_reduction/data/" + data_type + "/"
        navigation_dict = {key: None for key in param_dict.keys()}
        navigation_dict["nK"] = list(subdirs(path_prefix))
        navigation_dict["nL"] = list(subdirs(path_prefix + navigation_dict["nK"][0]))
        all_filenames = find_files(path_prefix)
        print(len(all_filenames))
        filenames = remove_extra_files(all_filenames, param_dict)
        print(len(filenames))

        alph_set, nK_set, nL_set, data_arr, param_arr = get_entries(filenames, entry)
        alph_ref = np.sort(list(alph_set))
        nK_ref = np.sort(list(nK_set))
        nL_ref = np.sort(list(nL_set))
        output_arr = [[[[] for _ in range(len(alph_ref))] for _ in range(len(nL_ref))] for _ in range(len(nK_ref))]

        for idx, params in list(enumerate(param_arr)):

            ([idx_nK],) = np.where(nK_ref == params[0])
            ([idx_nL],) = np.where(nL_ref == params[1])
            ([idx_alph],) = np.where(alph_ref == params[2])

            output_arr[idx_nK][idx_nL][idx_alph] = data_arr[idx]
            
        output_arr = np.squeeze(np.array(output_arr))
        output_name = entry + "_nK=" + nK + "_nL=" + nL + "_alph=" + alph + "_.npy"
        print(output_name)
        np.save(output_name, output_arr)

        return output_arr, nK_ref, nL_ref, alph_ref


class DataWriter(object):

    def __init__(self):
        self.none = None

    @staticmethod
    def create_model_workbook(path, data_dict):

        data_keys = data_dict.keys()
        wb = Workbook()
        wb.create_sheet(title="params")
        wb.create_sheet(title="eval")
        for key in data_keys:
            wb.create_sheet(title=key)
        wb.save(path)

    @staticmethod
    def write_model_workbook(path, param_dict, eval_dict, data_dict):

        def write_to_excel(data, ws):
            if isinstance(data, list) or type(data).__name__ == 'ndarray':
                df = pd.DataFrame(data)
                for r in dataframe_to_rows(df, index=False, header=False):
                    ws.append(r)
            else:
                ws.append([data])

        print("writing ...")

        wb = pxl.load_workbook(path)

        param_keys = param_dict.keys()
        eval_keys = eval_dict.keys()
        data_keys = data_dict.keys()

        active_ws = wb["params"]
        idx = wb.sheetnames.index("params")
        wb.remove(active_ws)
        wb.create_sheet("params", idx)
        active_ws = wb["params"]

        for key in param_keys:
            data = param_dict[key]
            write_to_excel([data], active_ws)
        active_ws.insert_cols(0)
        i = 1
        for key in param_keys:
            active_ws['A' + str(i)] = key
            i += 1

        active_ws = wb["eval"]
        idx = wb.sheetnames.index("eval")
        wb.remove(active_ws)
        wb.create_sheet("eval", idx)
        active_ws = wb["eval"]

        for key in eval_keys:
            data = eval_dict[key]
            write_to_excel([data], active_ws)
        active_ws.insert_cols(0)
        i = 1
        for key in eval_keys:
            active_ws['A' + str(i)] = key
            i += 1

        for key in data_keys:
            active_ws = wb[key]
            idx = wb.sheetnames.index(key)
            wb.remove(active_ws)
            wb.create_sheet(key, idx)
            active_ws = wb[key]
            data = data_dict[key]
            write_to_excel(data, active_ws)

        wb.save(path)

    @staticmethod
    def create_pred_workbook(path, pred_dict):
        pred_keys = pred_dict.keys()
        wb = Workbook()
        for key in pred_keys:
            wb.create_sheet(title=key)
        wb.save(path)

    @staticmethod
    def write_pred_workbook(path, pred_dict):

        def write_to_excel(data, ws):
            if isinstance(data, list) or type(data).__name__ == 'ndarray':
                df = pd.DataFrame(data)
                for r in dataframe_to_rows(df, index=False, header=False):
                    ws.append(r)
            else:
                ws.append([data])

        print("writing ...")

        wb = pxl.load_workbook(path)

        pred_keys = pred_dict.keys()
        for key in pred_keys:
            active_ws = wb[key]
            idx = wb.sheetnames.index(key)
            wb.remove(active_ws)
            wb.create_sheet(key, idx)
            active_ws = wb[key]
            data = pred_dict[key]
            write_to_excel(data, active_ws)

        wb.save(path)


class Predict_From_Model(object):

    def __init__(self):

        pred_dict = {
            "Q": None,
            "Z": None,
            "Zfr": None,
            "C": None,
            "KL": None,
            "xs_test": None

        }

        self.model_param_dict = None
        self.model_data_dict = None
        self.model_eval_dict = None
        self.pred_dict = pred_dict

    def load_model(self, path_to_model):

        loader = DataLoader()
        self.model_param_dict, self.model_eval_dict, self.model_data_dict = loader.load_model(path_to_model)

    def init_free_components(self, Z_train, Z_pred, validate):

        nPC = self.model_param_dict["nPC"]
        nL = self.model_param_dict["nL"]
        if validate is False:
            idx_test = self.model_data_dict["idx_test"]
        if validate is True:
            idx_test = self.model_data_dict["idx_val"]

        '''
        Zfr are the Z scores of the free components
        z2 are the Z scores of the components constrained by PCA
        '''

        muZ = np.mean(Z_train, axis=0).T
        muZ = muZ[:, np.newaxis]
        z2 = Z_pred[:, :nPC]  # PCA componetns when predicting # THESE ARE NOT FROM TRAINING # unknown for new cow
        mu2 = muZ[:nPC]
        mu1 = muZ[nPC:]
        covZ = np.cov(Z_train.T)  # Z is from training
        cov11 = covZ[nPC:, nPC:]
        cov12 = covZ[nPC:, :nPC]
        cov21 = cov12.T
        cov22 = covZ[:nPC, :nPC]

        # MEANS AND COVARIANCES ARE FROM TRAINING PHASE, z1 IS FROM THE TESTING PHASE

        sigbar = cov11 - cov12 @ np.linalg.inv(cov22) @ cov21
        nTest = len(idx_test)
        Zfr = np.zeros((nTest, nL))
        for i in range(nTest):
#             print(Z_train.shape)
#             print(nTest)
#             print("mu2: ", mu2.shape)
#             print("cov22: ", cov22.shape)
            z2_train = Z_pred[i, :nPC]
            z2_train = z2_train[:, np.newaxis]

            mubar = mu1 + cov12 @ np.linalg.inv(cov22) @ (mu2 - z2_train)
            Zfr[i, :] = np.random.multivariate_normal(mubar.ravel(), sigbar)

        Z = np.hstack((Z_pred, Zfr))

        return Zfr, Z

    def predict_microbiome(self, model_param_dict, model_data_dict, validate=False):

        pred_dict = self.pred_dict
        self.model_param_dict = model_param_dict
        self.model_data_dict = model_data_dict

        if validate is False:
            test_xs = model_data_dict["xs_test"]
            metz_test = model_data_dict["metz_test"]
            idx_test =model_data_dict["idx_test"]

        if validate is True:
            test_xs = model_data_dict["xs_val"]
            metz_test = model_data_dict["metz_val"]
            idx_test = model_data_dict["idx_val"]

        C = model_data_dict["C"]
        Z_train = model_data_dict["Z"]
        thet = model_data_dict["thet"]

        # Performing PCA reduction
        Z_pred = metz_test @ np.linalg.pinv(C)
        if model_param_dict["nL"] != 0:
            self.pred_dict["Zfr"], self.pred_dict["Z"] = self.init_free_components(Z_train, Z_pred, validate)
            Z = pred_dict["Z"]
        else:
            Z = Z_pred

        Q = np.exp(-Z @ thet)
        Q[np.isnan(Q)] = 0
        norms = Q.sum(1)
        Q = np.divide(Q.T, norms, where=norms != 0).T
        pred_dict["Q"] = Q

        KL = KL_divergence(test_xs, Q)/len(idx_test)

        pred_dict["xs_test"] = test_xs
        pred_dict["KL"] = KL

        self.pred_dict = pred_dict
        
    def predict_metadata(self, model_param_dict, model_data_dict, validate=False):
        
        pred_dict = self.pred_dict
        self.model_param_dict = model_param_dict
        self.model_data_dict = model_data_dict

        if validate is False:
            test_xs = model_data_dict["xs_test"]
            metz_test = model_data_dict["metz_test"]
        if validate is True:
            test_xs = model_data_dict["xs_val"]
            metz_test = model_data_dict["metz_val"]
            
        C = model_data_dict["C"]
        Z_train = model_data_dict["Z"]
        thet = model_data_dict["thet"]
        Q = model_data_dict["Q"]
        nPC = model_param_dict["nPC"]
        
        Z_pred = -np.log(Q) @ np.linalg.pinv(thet)
        Z_restricted = Z_pred[0:nPC, :]
        metz_pred = Z_restricted @ C
        
        KL = KL_divergence(metz_test, metz_pred)
        
        pred_dict["metz"] = metz_pred
        pred_dict["KL"] = KL
        
        self.pred_dict = pred_dict

    def finalize(self):
        return self.pred_dict
